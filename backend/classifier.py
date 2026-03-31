import os
import re
import csv
import json
import tempfile
from datetime import datetime

import anthropic
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import pdfplumber
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False

# ---------------------------------------------------------------------------
# ANTHROPIC CLIENT
# ---------------------------------------------------------------------------

client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY", ""))

# ---------------------------------------------------------------------------
# PARTE 1 — FILTRAR TRANSACCIONES
# ---------------------------------------------------------------------------

EXCLUDE_KEYWORDS = [
    "payment thank you", "thank you for your payment",
    "autopay", "auto pay", "credit card payment", "credit card pymt",
    "online payment", "mobile payment", "electronic payment",
    "minimum payment", "automatic payment",
]

EXCLUDE_DESC_IF_POSITIVE = [
    "payment", "thank you", "credit", "transfer", "refund",
    "deposit", "paycheck", "direct dep", "payroll deposit",
    "zelle", "venmo", "cashapp", "cash app",
]

EXCLUDE_TYPES = {"payment", "credit", "transfer", "credit card payment", "credit card pymt"}


def is_excluded(tx: dict) -> bool:
    desc = tx["description"].lower()
    amount = tx["amount"]
    tx_type = tx.get("chase_category", "").lower().strip()

    if tx_type in EXCLUDE_TYPES:
        return True

    for kw in EXCLUDE_KEYWORDS:
        if kw in desc:
            return True

    if amount > 0:
        for kw in EXCLUDE_DESC_IF_POSITIVE:
            if kw in desc:
                return True

    return False


def filter_transactions(transactions: list) -> list:
    filtered = []
    for tx in transactions:
        if not is_excluded(tx):
            tx = dict(tx)
            tx["amount"] = abs(tx["amount"])
            filtered.append(tx)
    return filtered


# ---------------------------------------------------------------------------
# PARTE 2 — CLASIFICACIÓN CON IA
# ---------------------------------------------------------------------------

CHUNK_SIZE = 50  # Max transactions per AI call to stay within token limits


def _call_ai_chunk(transactions: list, industry: str, id_offset: int = 0) -> list:
    """Call Claude API for a chunk of transactions. id_offset adjusts the displayed IDs."""
    tx_list = "\n".join(
        [f"{id_offset + i + 1}. {t['description']} | ${t['amount']:.2f}"
         for i, t in enumerate(transactions)]
    )

    prompt = f"""Eres un experto en impuestos de negocios en USA.
Clasifica cada gasto según el IRS Schedule C para un negocio de industria: {industry}

TRANSACCIONES:
{tx_list}

REGLAS IMPORTANTES — Los bancos truncan nombres, busca el patrón:
- Airlines (Southwest/WN, Delta/DL, United/UA, American/AA, Spirit, Frontier, JetBlue, ALLEGIANT) = Travel | Schedule C - Line 24a
- Hotels (Marriott, Hilton, Hyatt, Airbnb, Holiday Inn, Best Western, Comfort Inn, Hampton Inn, VRBO) = Travel | Schedule C - Line 24a
- Uber, Lyft, LYFT *RIDE, UBER *TRIP = Travel | Schedule C - Line 24a
- Gas stations (Shell, BP, Chevron, Exxon, Mobil, Kwik Trip, Pilot, Loves, Flying J, Circle K, Speedway, Racetrac, Wawa, QT, Valero, Sunoco, Murphy, GetGo) = Fuel | Schedule C - Line 9
- Restaurants/food (McDonald's, MCD, Starbucks, SBUX, Subway, Chipotle, Taco Bell, Domino's, Pizza, KFC, Popeyes, Chick-fil-A, Sonic, Arby's, Wendy's, Panera, Dunkin, IHOP, Denny's, Waffle House, Applebee's, any TST* or SQ * at food place) = Meals (50% Deductible) | Schedule C - Line 24b
- Food delivery (Uber Eats, UBER* EATS, DoorDash, DD *, GrubHub, Postmates, Instacart food) = Meals (50% Deductible) | Schedule C - Line 24b
- Phone/Internet (AT&T, Verizon, T-Mobile, Comcast, Xfinity, Spectrum, Cox, CenturyLink, Cricket) = Utilities | Schedule C - Line 25
- Software (Microsoft, MSFT, Adobe, Google, AWS, Amazon Web, Zoom, Slack, QuickBooks, Intuit, Dropbox, Shopify, Squarespace, Wix, HubSpot, Mailchimp, Canva) = Software & Subscriptions | Schedule C - Line 27a
- Payroll (Gusto, ADP, Paychex, Rippling, Zenefits) = Wages & Salaries | Schedule C - Line 26
- Insurance (Geico, Progressive, State Farm, Allstate, Nationwide, Liberty Mutual, Travelers, Farmers, USAA, any *INSURANCE*) = Insurance | Schedule C - Line 15
- Rent/Lease/Storage (Public Storage, Extra Space, U-Haul storage, any RENT or LEASE) = Rent & Lease | Schedule C - Line 20b
- Legal/CPA/Accounting/Consulting = Legal & Professional | Schedule C - Line 17
- Advertising (Facebook, META *, Google Ads, Instagram, LinkedIn, Yelp, any *ADS*) = Advertising | Schedule C - Line 8
- Auto repair (AutoZone, O'Reilly, OREILLY, Advance Auto, Napa Auto, Jiffy Lube, Firestone, Midas, Pep Boys, Valvoline, Meineke, Maaco, any *AUTO REPAIR*) = Car & Truck Expenses | Schedule C - Line 9
- Tolls/Parking (IPASS, EZPass, BESTPASS, ParkWhiz, SpotHero, any PARKING, any TOLL) = Tolls & Parking | Schedule C - Line 9
- Bank fees, interest, Stripe, PayPal, Square (SQ *), Clover, Toast (TST*) fees = Bank & Processing Fees | Schedule C - Line 27a
- Amazon (AMZN MKTP, AMAZON.COM), Office Depot, Staples, Home Depot, Lowe's, Costco Business = Supplies | Schedule C - Line 22
- Walmart (WM SUPERCENTER, WALMART), Target, Costco personal, Sam's Club personal = Personal (Non-Deductible) | N/A
- Netflix, Hulu, Spotify, Disney+, HBO, Apple TV = Personal (Non-Deductible) | N/A
- Si no encaja en ninguna categoría de negocio, clasifica como la más cercana (NO dejes sin categorizar)

Responde SOLO con un JSON array, sin texto extra ni markdown:
[
  {{"id": {id_offset + 1}, "category": "Travel", "irs_line": "Schedule C - Line 24a", "deductible": "YES", "confidence": "HIGH"}},
  {{"id": {id_offset + 2}, "category": "Meals (50% Deductible)", "irs_line": "Schedule C - Line 24b", "deductible": "50%", "confidence": "HIGH"}}
]

Valores para deductible: "YES", "NO", "50%"
Valores para confidence: "HIGH", "MEDIUM", "LOW"

Clasifica los {len(transactions)} gastos. TODOS deben tener una categoría asignada."""

    response = client.messages.create(
        model="claude-opus-4-6",
        max_tokens=8096,
        messages=[{"role": "user", "content": prompt}]
    )

    raw = response.content[0].text.strip()
    raw = re.sub(r"^```[a-zA-Z]*\s*", "", raw)
    raw = re.sub(r"\s*```\s*$", "", raw)
    raw = raw.strip()
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        match = re.search(r"\[.*\]", raw, re.DOTALL)
        if match:
            return json.loads(match.group())
        raise


def classify_batch_with_ai(transactions: list, industry: str) -> list:
    if not transactions:
        return []

    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        raise ValueError(
            "ANTHROPIC_API_KEY no está configurada. "
            "Configura la variable de entorno para habilitar la clasificación con IA."
        )

    all_results = []
    for i in range(0, len(transactions), CHUNK_SIZE):
        chunk = transactions[i:i + CHUNK_SIZE]
        chunk_results = _call_ai_chunk(chunk, industry, id_offset=i)
        all_results.extend(chunk_results)

    return all_results


def classify_fallback(desc: str) -> dict:
    # Real bank description patterns: truncated names, codes, prefixes like SQ*, TST*, AMZN
    RULES = [
        ("Fuel", "Schedule C - Line 9", "YES",
         ["shell", "bp ", "chevron", "exxon", "mobil", "fuel", "gasoline", "gas station",
          "circle k", "pilot trav", "loves trav", "flying j", "speedway", "kwik trip",
          "marathon", "sunoco", "valero", "quiktrip", "racetrac", "wawa", "murphy",
          "getgo", "casey", "kum & go", "holiday station", "thorntons", "sheetz",
          "bucees", "buc-ee", "pump", "76 oil", "76gas"]),
        ("Car & Truck Expenses", "Schedule C - Line 9", "YES",
         ["autozone", "oreilly", "o'reilly", "advance auto", "napa auto", "jiffy lube",
          "firestone", "midas", "pep boys", "oil change", "tire", "brake", "auto repair",
          "car wash", "truck repair", "mechanic", "valvoline", "meineke", "maaco",
          "monro", "christian brothers", "take 5 oil", "jiffy", "tires plus",
          "discount tire", "ntb ", "sears auto", "goodyear", "bridgestone"]),
        ("Meals (50% Deductible)", "Schedule C - Line 24b", "50%",
         ["mcdonald", "mcd ", "burger king", "bk ", "wendy", "subway", "taco bell",
          "chipotle", "panera", "domino", "pizza hut", "little caesar", "papa john",
          "kfc ", "popeyes", "chick-fil-a", "cfa ", "sonic ", "arby", "dairy queen",
          "dq ", "dunkin", "starbucks", "sbux", "tim horton", "ihop", "denny",
          "waffle house", "cracker barrel", "applebee", "chili", "olive garden",
          "red lobster", "outback", "texas roadhouse", "grubhub", "doordash", "dd ",
          "ubereats", "uber* eat", "postmates", "instacart", "restaurant", "cafe",
          "diner", "bakery", "sushi", "tst*", "sq *coffee", "sq *cafe", "sq *bakery",
          "sq *restaurant", "sq *bar", "sq *grill", "sq *kitchen", "sq *food",
          "jersey mike", "jimmy john", "five guys", "shake shack", "in-n-out",
          "whataburger", "culver", "raising cane", "wingstop", "zaxby", "cook out",
          "bojangles", "hardee", "carl's jr", "jack in the box"]),
        ("Travel", "Schedule C - Line 24a", "YES",
         ["hotel", "motel", "inn ", "lodging", "marriott", "hilton", "hyatt", "airbnb",
          "holiday inn", "best western", "comfort inn", "hampton inn", "courtyard",
          "residence inn", "fairfield", "springhill", "towneplace", "embassy suites",
          "doubletree", "aloft ", "element ", "westin", "sheraton", "wyndham",
          "la quinta", "days inn", "super 8", "extended stay", "vrbo",
          "southwest air", "delta air", "united air", "american air", "spirit air",
          "frontier air", "jetblue", "allegiant", "alaska air", "hawaiian air",
          "sun country", "flight", "airfare", "airline", "air ticket",
          "enterprise rent", "hertz", "avis ", "budget car", "national car", "alamo",
          "thrifty car", "dollar rent", "fox rent", "silvercar", "turo ",
          "lyft *ride", "uber *trip", "uber trip", "lyft ride",
          "amtrak", "greyhound", "megabus"]),
        ("Tolls & Parking", "Schedule C - Line 9", "YES",
         ["parking", "toll ", "ezpass", "ez pass", "ipass", "i-pass", "bestpass",
          "parkwhiz", "spothero", "turnpike", "expressway", "tollway", "laparking",
          "park mobile", "parkmobile", "paybyphone", "meterfeeder"]),
        ("Insurance", "Schedule C - Line 15", "YES",
         ["insurance", "insur", "geico", "progressive", "state farm", "allstate",
          "nationwide", "liberty mutual", "travelers", "farmers ins", "usaa",
          "hartford", "chubb", "aig ", "workers comp", "general liability"]),
        ("Wages & Salaries", "Schedule C - Line 26", "YES",
         ["payroll", "gusto", "adp ", "paychex", "rippling", "zenefits", "wages",
          "salary", "direct deposit payroll", "bamboohr"]),
        ("Utilities", "Schedule C - Line 25", "YES",
         ["verizon", "at&t", "att ", "t-mobile", "tmobile", "comcast", "xfinity",
          "spectrum", "cox comm", "centurylink", "lumen", "windstream", "frontier comm",
          "cricket wireless", "boost mobile", "metro pcs", "metropcs",
          "electric", "electricity", "water bill", "utility", "internet bill",
          "duke energy", "con edison", "pge ", "pg&e", "dte energy", "dominion",
          "xcel energy", "westar", "evergy", "national grid"]),
        ("Rent & Lease", "Schedule C - Line 20b", "YES",
         ["rent ", "lease ", "storage unit", "office space", "public storage",
          "extra space", "life storage", "cubesmart", "u-haul storage", "uhaul storage",
          "simply storage", "warehouse", "coworking", "wework", "regus", "spaces "]),
        ("Legal & Professional", "Schedule C - Line 17", "YES",
         ["attorney", "lawyer", "legal ", "accountant", "cpa ", "bookkeeping",
          "consulting", "notary", "paralegal", "enrolled agent", "tax prep",
          "h&r block", "jackson hewitt", "legalzoom"]),
        ("Software & Subscriptions", "Schedule C - Line 27a", "YES",
         ["quickbooks", "intuit", "microsoft", "msft", "google workspace", "gsuite",
          "adobe ", "dropbox", "slack ", "zoom ", "shopify", "squarespace", "wix ",
          "aws ", "amazon web", "digitalocean", "linode", "heroku", "github",
          "hubspot", "mailchimp", "klaviyo", "constantcontact", "hootsuite",
          "buffer ", "canva ", "figma ", "notion ", "asana ", "monday.com",
          "trello", "basecamp", "freshbooks", "xero ", "wave acc", "sage ",
          "subscription", "saas", "software lic", "app store", "google play",
          "apple.com/bill", "icloud", "spotify business", "pandora business"]),
        ("Advertising", "Schedule C - Line 8", "YES",
         ["facebook", "meta ", "google ads", "instagram", "linkedin ads", "twitter ads",
          "snapchat ads", "tiktok ads", "yelp ", "nextdoor ads", "pinterest ads",
          "advertising", "marketing", "promotion", "signage", "print ad",
          "radio ad", "tv ad", "billboard", "seo ", "sem ", "ppc "]),
        ("Supplies", "Schedule C - Line 22", "YES",
         ["office depot", "officemax", "staples", "home depot", "homedepot",
          "lowe's", "lowes", "menards", "ace hardware", "true value",
          "amzn mktp", "amazon.com", "amazon mktpl", "uline", "grainger",
          "fastenal", "harbor freight", "northern tool", "supplies", "tools "]),
        ("Bank & Processing Fees", "Schedule C - Line 27a", "YES",
         ["bank fee", "monthly fee", "stripe", "paypal", "sq *", "square inc",
          "clover ", "toast ", "tst*", "processing fee", "merchant fee",
          "service charge", "overdraft", "nsf fee", "wire fee", "finance charge",
          "interest charge", "late fee"]),
        ("Personal (Non-Deductible)", "N/A - Not Deductible", "NO",
         ["wm supercenter", "walmart", "wal-mart", "target ", "costco", "sam's club",
          "samsclub", "kroger", "publix", "whole foods", "wholefds", "trader joe",
          "safeway", "aldi ", "heb ", "meijer", "giant ", "stop & shop",
          "netflix", "hulu ", "spotify", "disney+", "hbo ", "amazon prime video",
          "apple tv", "peacock", "paramount+", "gym", "fitness", "planet fitness",
          "anytime fitness", "la fitness", "ymca", "salon", "spa ", "nail ",
          "hair cut", "barber", "casino", "gambling", "lottery"]),
    ]

    text = desc.lower()
    best = None
    best_score = 0

    for category, irs_line, deductible, keywords in RULES:
        score = sum(1 for kw in keywords if kw in text)
        if score > best_score:
            best_score = score
            best = (category, irs_line, deductible)

    if best:
        return {
            "category": best[0],
            "irs_line": best[1],
            "deductible": best[2],
            "confidence": "MEDIUM" if best_score >= 2 else "LOW",
        }

    # Last resort: tag as supplies if it looks like a business purchase
    return {
        "category": "Other Business Expense",
        "irs_line": "Schedule C - Line 27a",
        "deductible": "YES",
        "confidence": "LOW",
    }


def classify_all(transactions: list, industry: str) -> list:
    if not transactions:
        return []

    try:
        ai_results = classify_batch_with_ai(transactions, industry)
        ai_map = {r["id"]: r for r in ai_results}

        classified = []
        for i, tx in enumerate(transactions):
            tx = dict(tx)
            ai = ai_map.get(i + 1)
            if ai:
                tx["category"] = ai.get("category", "Uncategorized")
                tx["irs_line"] = ai.get("irs_line", "Review Required")
                tx["deductible"] = ai.get("deductible", "NO")
                tx["confidence"] = ai.get("confidence", "LOW")
            else:
                fb = classify_fallback(tx["description"])
                tx.update(fb)
            classified.append(tx)
        return classified

    except ValueError as e:
        # Missing API key or config error — surface to user
        raise
    except Exception as e:
        # Transient error (network, timeout, etc.) — use fallback silently
        print(f"[AI classification failed, using keyword fallback]: {e}")
        classified = []
        for tx in transactions:
            tx = dict(tx)
            fb = classify_fallback(tx["description"])
            tx.update(fb)
            classified.append(tx)
        return classified


# ---------------------------------------------------------------------------
# PARTE 3 — LEER ARCHIVOS (UNIVERSAL)
# ---------------------------------------------------------------------------

def normalize_rows(rows: list, headers: list) -> list:
    DATE_HINTS = ["date", "fecha", "posting", "trans date", "transaction date", "post date"]
    DESC_HINTS = ["desc", "description", "memo", "details", "payee", "merchant",
                  "name", "narration", "particulars", "transaction"]
    AMT_HINTS  = ["amount", "amt", "debit", "charge", "monto", "value", "sum", "total", "credit"]
    CAT_HINTS  = ["category", "type", "cat", "transaction type", "chase category"]
    TYPE_HINTS = ["type", "transaction type", "trans type"]

    def best_col(hints):
        for h in hints:
            for i, header in enumerate(headers):
                if h in header.lower():
                    return i
        return None

    date_col = best_col(DATE_HINTS)
    desc_col = best_col(DESC_HINTS)
    amt_col  = best_col(AMT_HINTS)
    cat_col  = best_col(CAT_HINTS)
    type_col = best_col(TYPE_HINTS)

    if desc_col is None:
        desc_col = min(1, len(headers) - 1)
    if amt_col is None:
        amt_col = min(2, len(headers) - 1)

    normalized = []
    for row in rows:
        if not row:
            continue
        date    = str(row[date_col]).strip() if date_col is not None and date_col < len(row) else ""
        desc    = str(row[desc_col]).strip() if desc_col is not None and desc_col < len(row) else ""
        cat     = str(row[cat_col]).strip()  if cat_col is not None and cat_col < len(row) else ""
        tx_type = str(row[type_col]).strip() if type_col is not None and type_col < len(row) else ""

        raw_amt = str(row[amt_col]).strip() if amt_col is not None and amt_col < len(row) else "0"
        raw_amt = re.sub(r"[^\d.\-]", "", raw_amt)
        try:
            amount = float(raw_amt) if raw_amt not in ("", "-") else 0.0
        except ValueError:
            amount = 0.0

        if not desc or desc.lower() in ("description", "memo", "details", "transaction", "name", "payee"):
            continue
        if amount == 0.0:
            continue

        normalized.append({
            "date": date,
            "description": desc,
            "amount": amount,
            "chase_category": cat if cat else tx_type,
        })

    return normalized


def read_csv(file_path: str) -> list:
    with open(file_path, newline="", encoding="utf-8-sig", errors="replace") as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(f, dialect)
        rows = list(reader)

    if not rows:
        return []

    header_idx = 0
    for i, row in enumerate(rows):
        if any(cell.strip() for cell in row):
            header_idx = i
            break

    headers = [c.strip() for c in rows[header_idx]]
    return normalize_rows(rows[header_idx + 1:], headers)


def read_excel(file_path: str) -> list:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append([str(c) if c is not None else "" for c in row])
    wb.close()

    if not all_rows:
        return []

    header_idx = 0
    for i, row in enumerate(all_rows):
        if any(str(c).strip() for c in row):
            header_idx = i
            break

    headers = [str(c).strip() for c in all_rows[header_idx]]
    return normalize_rows(all_rows[header_idx + 1:], headers)


def read_pdf(file_path: str) -> list:
    if not PDF_SUPPORT:
        raise RuntimeError("pdfplumber not installed.")

    TX_PATTERN = re.compile(
        r"(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})\s+(.+?)\s+([\-]?\$?[\d,]+\.\d{2})"
    )
    transactions = []
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.splitlines():
                m = TX_PATTERN.search(line)
                if m:
                    date, desc, raw_amt = m.group(1), m.group(2).strip(), m.group(3)
                    raw_amt = re.sub(r"[^\d.\-]", "", raw_amt)
                    try:
                        amount = float(raw_amt)
                    except ValueError:
                        amount = 0.0
                    transactions.append({
                        "date": date,
                        "description": desc,
                        "amount": amount,
                        "chase_category": "",
                    })
    return transactions


# ---------------------------------------------------------------------------
# EXCEL BUILDER
# ---------------------------------------------------------------------------

DARK_BLUE   = "1F3864"
LIGHT_BLUE  = "D6E4F0"
GREEN       = "C6EFCE"
GREEN_FONT  = "276221"
RED         = "FFCCCC"
RED_FONT    = "9C0006"
YELLOW      = "FFEB9C"
YELLOW_FONT = "9C6500"
WHITE       = "FFFFFF"
GRAY        = "F2F2F2"


def _header_fill(color=DARK_BLUE):
    return PatternFill("solid", fgColor=color)

def _cell_fill(color):
    return PatternFill("solid", fgColor=color)

def _thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def _header_font(color=WHITE, bold=True, size=11):
    return Font(name="Calibri", bold=bold, color=color, size=size)

def _normal_font(color="000000", bold=False, size=10):
    return Font(name="Calibri", bold=bold, color=color, size=size)

def _set_col_width(ws, col_idx, width):
    ws.column_dimensions[get_column_letter(col_idx)].width = width


def _row_colors(tx: dict):
    ded = tx.get("deductible", "NO").upper()
    if ded == "YES":
        return GREEN, GREEN_FONT
    elif ded == "50%":
        return YELLOW, YELLOW_FONT
    elif tx.get("category", "") == "Uncategorized":
        return YELLOW, YELLOW_FONT
    else:
        return RED, RED_FONT


def _build_all_transactions(wb, transactions: list):
    ws = wb.create_sheet("All Transactions")
    headers = ["Date", "Description", "Amount", "Category", "Schedule C Line", "Deductible", "Confidence"]
    col_widths = [14, 48, 14, 30, 35, 12, 12]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _header_fill(DARK_BLUE)
        cell.font = _header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
        _set_col_width(ws, col, w)
    ws.row_dimensions[1].height = 30

    for row_idx, tx in enumerate(transactions, start=2):
        fill_color, font_color = _row_colors(tx)
        ded = tx.get("deductible", "NO").upper()
        ded_display = "50%" if ded == "50%" else ("Yes" if ded == "YES" else "No")

        values = [
            tx.get("date", ""),
            tx.get("description", ""),
            tx.get("amount", 0.0),
            tx.get("category", "Uncategorized"),
            tx.get("irs_line", "Review Required"),
            ded_display,
            tx.get("confidence", "LOW").title(),
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = _cell_fill(fill_color)
            cell.font = _normal_font(color=font_color)
            cell.border = _thin_border()
            if col == 3:
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col in (1, 6, 7):
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:G1"
    return ws


def _build_summary(wb, transactions: list):
    ws = wb.create_sheet("Summary by Category")
    summary = {}
    for tx in transactions:
        cat = tx.get("category", "Uncategorized")
        if cat not in summary:
            summary[cat] = {
                "category": cat,
                "irs_line": tx.get("irs_line", "Review Required"),
                "deductible": tx.get("deductible", "NO"),
                "count": 0,
                "total": 0.0,
            }
        summary[cat]["count"] += 1
        summary[cat]["total"] += abs(tx.get("amount", 0.0))

    rows_data = sorted(
        summary.values(),
        key=lambda x: (x["deductible"] == "NO", x["category"] == "Uncategorized", -x["total"])
    )

    headers = ["Category", "Schedule C Line", "Deductible", "# Transactions", "Total Amount"]
    col_widths = [32, 38, 12, 16, 18]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _header_fill(DARK_BLUE)
        cell.font = _header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()
        _set_col_width(ws, col, w)
    ws.row_dimensions[1].height = 28

    total_deductible = 0.0
    total_non_deductible = 0.0

    for row_idx, item in enumerate(rows_data, start=2):
        ded = item["deductible"].upper()
        if ded == "YES":
            fc, ff = GREEN, GREEN_FONT
        elif ded == "50%":
            fc, ff = YELLOW, YELLOW_FONT
        elif item["category"] == "Uncategorized":
            fc, ff = YELLOW, YELLOW_FONT
        else:
            fc, ff = RED, RED_FONT

        ded_display = "50%" if ded == "50%" else ("Yes" if ded == "YES" else "No")
        vals = [item["category"], item["irs_line"], ded_display, item["count"], item["total"]]
        for col, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = _cell_fill(fc)
            cell.font = _normal_font(color=ff)
            cell.border = _thin_border()
            if col == 5:
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col in (3, 4):
                cell.alignment = Alignment(horizontal="center")

        if ded == "YES":
            total_deductible += item["total"]
        elif ded == "50%":
            total_deductible += item["total"] * 0.5
            total_non_deductible += item["total"] * 0.5
        else:
            total_non_deductible += item["total"]

    last = len(rows_data) + 2
    c = ws.cell(row=last, column=1, value="TOTAL DEDUCTIBLE EXPENSES")
    c.font = _header_font(color=GREEN_FONT)
    c.fill = _cell_fill(GREEN)
    c = ws.cell(row=last, column=5, value=total_deductible)
    c.number_format = '"$"#,##0.00'
    c.font = _header_font(color=GREEN_FONT)
    c.fill = _cell_fill(GREEN)

    c = ws.cell(row=last + 1, column=1, value="TOTAL NON-DEDUCTIBLE")
    c.font = _header_font(color=RED_FONT)
    c.fill = _cell_fill(RED)
    c = ws.cell(row=last + 1, column=5, value=total_non_deductible)
    c.number_format = '"$"#,##0.00'
    c.font = _header_font(color=RED_FONT)
    c.fill = _cell_fill(RED)

    ws.freeze_panes = "A2"
    return ws


IRS_NOTES = [
    ("Fuel", "Schedule C - Line 9",
     "Keep mileage log OR actual expense receipts. Cannot deduct both standard mileage and actual expenses."),
    ("Car & Truck Expenses", "Schedule C - Line 9",
     "Only the business-use percentage is deductible. Keep records of business vs personal use."),
    ("Meals (50% Deductible)", "Schedule C - Line 24b",
     "Only 50% of business meal expenses are deductible. Must have business purpose documented."),
    ("Travel", "Schedule C - Line 24a",
     "Travel must be business-related and away from your tax home. Keep receipts and business purpose."),
    ("Tolls & Parking", "Schedule C - Line 9",
     "Business tolls and parking are deductible as car and truck expenses."),
    ("Insurance", "Schedule C - Line 15",
     "Business insurance premiums are fully deductible. Health insurance for self-employed on Schedule 1."),
    ("Wages & Salaries", "Schedule C - Line 26",
     "Employee wages fully deductible. 1099 contractors reported on Schedule C as well."),
    ("Legal & Professional", "Schedule C - Line 17",
     "Legal, accounting, and consulting fees directly related to business are fully deductible."),
    ("Advertising", "Schedule C - Line 8",
     "All ordinary and necessary advertising costs are fully deductible."),
    ("Rent & Lease", "Schedule C - Line 20b",
     "Rent for business property fully deductible. Vehicle lease: only business-use portion deductible."),
    ("Utilities", "Schedule C - Line 25",
     "Business utilities are fully deductible. Only the business-use percentage for mixed use."),
    ("Taxes & Licenses", "Schedule C - Line 23",
     "Business licenses, IFTA, and 2290 heavy vehicle use tax are deductible."),
    ("Software & Subscriptions", "Schedule C - Line 27a",
     "Business software subscriptions are fully deductible in the year paid."),
    ("Bank & Processing Fees", "Schedule C - Line 27a",
     "Bank fees and merchant processing fees for business accounts are fully deductible."),
    ("Supplies", "Schedule C - Line 22",
     "Supplies consumed or used during the tax year are deductible. Keep all receipts."),
    ("Personal (Non-Deductible)", "N/A",
     "Personal expenses are NOT deductible on Schedule C. Keep business and personal accounts separate."),
    ("Uncategorized", "Review Required",
     "These transactions need manual review. Categorize or consult your tax professional."),
]


def _build_irs_notes(wb, company_name, year, industry, entity, notes=""):
    ws = wb.create_sheet("IRS Notes")
    ws.merge_cells("A1:E1")
    title = ws.cell(row=1, column=1, value=f"IRS Schedule C Tax Notes — {company_name} ({year})")
    title.fill = _header_fill(DARK_BLUE)
    title.font = Font(name="Calibri", bold=True, color=WHITE, size=14)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    meta = [("Company", company_name), ("Tax Year", year), ("Industry", industry),
            ("Entity Type", entity), ("Generated", datetime.today().strftime("%Y-%m-%d"))]
    for i, (k, v) in enumerate(meta, start=2):
        ws.cell(row=i, column=1, value=k).font = _normal_font(bold=True)
        ws.cell(row=i, column=2, value=v).font = _normal_font()

    extra_rows = 0
    if notes and notes.strip():
        notes_row = len(meta) + 2
        ws.merge_cells(f"A{notes_row}:E{notes_row}")
        lbl = ws.cell(row=notes_row, column=1, value="Client Notes / Important Expenses")
        lbl.fill = _header_fill("B45309")
        lbl.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
        lbl.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[notes_row].height = 22
        nv = notes_row + 1
        ws.merge_cells(f"A{nv}:E{nv}")
        nc = ws.cell(row=nv, column=1, value=notes.strip())
        nc.font = _normal_font()
        nc.alignment = Alignment(wrap_text=True, vertical="top")
        nc.fill = _cell_fill("FFFBEB")
        ws.row_dimensions[nv].height = max(60, len(notes.strip()) // 3)
        extra_rows = 2

    header_row = len(meta) + 3 + extra_rows
    h_cols = ["Category", "Schedule C Line", "IRS Notes / Rules"]
    for col, (h, w) in enumerate(zip(h_cols, [30, 35, 70]), start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.fill = _header_fill(DARK_BLUE)
        cell.font = _header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[header_row].height = 24

    for r_idx, (cat, line, note) in enumerate(IRS_NOTES, start=header_row + 1):
        fill_color = LIGHT_BLUE if r_idx % 2 == 0 else GRAY
        for col, val in enumerate([cat, line, note], start=1):
            cell = ws.cell(row=r_idx, column=col, value=val)
            cell.fill = _cell_fill(fill_color)
            cell.font = _normal_font()
            cell.border = _thin_border()
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r_idx].height = 40

    ws.freeze_panes = f"A{header_row + 1}"
    return ws


def build_excel(transactions: list, company_name: str, year: str, industry: str,
                entity: str, notes: str = "") -> str:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _build_all_transactions(wb, transactions)
    _build_summary(wb, transactions)
    _build_irs_notes(wb, company_name, year, industry, entity, notes)
    out_dir = tempfile.gettempdir()
    safe_name = re.sub(r"[^a-zA-Z0-9_\-]", "_", company_name)
    out_path = os.path.join(out_dir, f"{safe_name}_IRS_Categories_{year}.xlsx")
    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# ENTRY POINTS
# ---------------------------------------------------------------------------

def _read_file(file_path: str, file_ext: str) -> list:
    ext = file_ext.lower()
    if ext == "csv":
        return read_csv(file_path)
    elif ext in ("xlsx", "xls"):
        return read_excel(file_path)
    elif ext == "pdf":
        return read_pdf(file_path)
    else:
        raise ValueError(f"Unsupported file type: {ext}")


def process_file(file_path: str, file_ext: str, company_name: str,
                 year: str, industry: str, entity: str) -> str:
    raw = _read_file(file_path, file_ext)
    transactions = filter_transactions(raw)
    if not transactions:
        raise ValueError("No expense transactions found after filtering.")
    classified = classify_all(transactions, industry)
    return build_excel(classified, company_name, year, industry, entity)


def process_file_full(file_path: str, file_ext: str, company_name: str,
                      year: str, industry: str, entity: str, notes: str = "") -> tuple:
    raw = _read_file(file_path, file_ext)
    transactions = filter_transactions(raw)
    if not transactions:
        raise ValueError("No expense transactions found after filtering.")

    classified = classify_all(transactions, industry)

    total_expenses = round(sum(tx["amount"] for tx in classified), 2)
    cat_totals: dict = {}
    for tx in classified:
        cat = tx.get("category", "Uncategorized")
        cat_totals[cat] = cat_totals.get(cat, 0.0) + tx["amount"]

    categories = sorted(
        [{"category": c, "total": round(t, 2)} for c, t in cat_totals.items()],
        key=lambda x: -x["total"],
    )

    summary = {
        "total_income": 0.0,
        "total_expenses": total_expenses,
        "net": round(-total_expenses, 2),
        "categories": categories,
        "transaction_count": len(classified),
    }

    excel_path = build_excel(classified, company_name, year, industry, entity, notes)
    return excel_path, summary
